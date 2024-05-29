import json
import os
import pickle
from django.shortcuts import render, redirect
import base64
import cv2
from django.http import StreamingHttpResponse
import face_recognition
from .tasks import registerStudent
from .models import Student, Department, Division, TempFileTest, Classrooms
from .predict import predict
from .threading import CameraWidget
import datetime
from openpyxl import Workbook, load_workbook
from django.contrib.auth.decorators import login_required

s = {}

def home_redirect(request):
    return redirect('studentregister')

def gen_frames(model, dept ,div):
    camera = cv2.VideoCapture(1, cv2.CAP_DSHOW)
    global s
    s[dept+div] = []
    while True:
        _, frame = camera.read()
        s[dept+div].append([predict(frame, model), str(datetime.datetime.now().strftime("%H:%M:%S")), str(datetime.datetime.now().strftime("%m/%d/%Y"))])
        #s[dept+div].add(tuple(predict(frame, model)))
        #print(frame.shape)
        frame = cv2.imencode('.jpeg', frame)[1].tobytes()
        yield (b'--frame\r\n'b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

        if cv2.waitKey(1) == 13:
            camera.release()
            cv2.destroyAllWindows()
            break

@login_required(login_url="/student/loginrequired/")
def vidfeed(request, dept, div):
    for_path = Student.objects.all()[0]
    path_model = for_path.img1.path
    path_model = os.path.split(path_model)
    for _ in range(4):
        path_model = os.path.split(path_model[0])
    path_model = f"{path_model[0]}\\models\\{dept}\\{div}\\model.pkl"
    #t1= CameraWidget(dept, div, path_model)
    file = open(path_model, 'rb')
    model = pickle.load(file)
    file.close()
    #t1.start()
    #print(t1)
    return StreamingHttpResponse(gen_frames(model, dept , div),content_type='multipart/x-mixed-replace; boundary=frame')
    #return StreamingHttpResponse(t1.gen_frames(),content_type='multipart/x-mixed-replace; boundary=frame')

@login_required(login_url="/student/loginrequired/")
def test_page(request, dept, div):
    workbook_name = 'attendance_xl.xlsx'

    if request.method == "POST":

        if os.path.isfile(workbook_name):
            workbook = load_workbook(workbook_name)
        else:
            workbook = Workbook()
        
        if dept+div in workbook.sheetnames:
            worksheet = workbook[dept+div]
        else:
            worksheet = workbook.create_sheet(dept+div)

        attendances = s.get(dept+div, None)
        #print(attendances)

        dics = {}


        #[['2JR19CS041','2JR20CS413'], '21:18:13', '05/12/2023']

        for each_frame in attendances:
            for usn in each_frame[0]:
                dics[usn] = [each_frame[1]]


        for dic_usn in dics.keys():
            for cells in worksheet.iter_cols(min_col=1, max_col=1, values_only=True):
                for xl_usn in range(len(cells)):
                    if dic_usn == cells[xl_usn]:
                        worksheet['B'+str(xl_usn+1)].value = dics[dic_usn][0]
                        worksheet['C'+str(xl_usn+1)].value = worksheet['C'+str(xl_usn+1)].value + 1
                        break
                else:
                    endcell_index = worksheet.max_row + 1
                    end_cell = worksheet.cell(row=worksheet.max_row+1,column=1)
                    end_cell.value = dic_usn
                    worksheet['B'+str(endcell_index)].value = dics[dic_usn][0]
                    worksheet['C'+str(endcell_index)].value = 1

        cnt = 0
        for i in range(worksheet.max_row):
            if cnt == 0:
                cnt += 1
                continue
            if worksheet['D'+str(i+1)].value is not None:
                worksheet['D'+str(i+1)].value = worksheet['D'+str(i+1)].value + 1
            else:
                worksheet['D'+str(i+1)].value = 1
                    
        workbook.save(workbook_name)
        '''for each_frame in attendances:
            for usn in each_frame[0]:
                for cells in worksheet.iter_cols(min_col=1, max_col=1, values_only=True):
                    for xl_usn in range(len(cells)):

                        if worksheet['D'+str(xl_usn+1)] is not None:
                            worksheet['D'+str(xl_usn+1)].value = worksheet['D'+str(xl_usn+1)].value + 1
                        else:
                            worksheet['D'+str(xl_usn+1)].value = 1
                        
                        if cells[xl_usn] == usn:
                            if worksheet['C'+str(xl_usn+1)] is not None:
                                worksheet['C'+str(xl_usn+1)].value = worksheet['C'+str(xl_usn+1)].value + 1
                            else:
                                worksheet['C'+str(xl_usn+1)].value = 1

                    #print(usn, each_frame[1], each_frame[2])'''

        #workbook.save('attendance_xl.xlsx')
        
        '''
        workbook = Workbook()
        sheet = workbook.active
        col = 1
        global s
        attendances = s.get(dept+div)
        for attende in attendances:
            print(attende)
            if len(attende) > 0:
                sheet['A'+str(col)] = attende[0]
                sheet['B'+str(col)] = datetime.datetime.now().strftime('%H:%M:%S')
                col += 1
            else:
                pass
        workbook.save(filename="hello_world.xlsx")
        col = 1
        '''
        return redirect('success')
    return render(request, 'student/base.html', {"dept":dept, "div":div})

def home_page(request):
    return render(request, 'index.html')

def student_home(request):
    divs = Division.objects.all() 
    depts = Department.objects.all()
    context = {}
    context['dept'] = depts
    context['div'] = divs
    if request.method =="POST":
        imgs = json.loads(request.POST['images'])
        name = request.POST['name']
        usn = request.POST['usn']
        dept = Department.objects.get(name=request.POST['dept'])
        div = Division.objects.get(div_name=request.POST['div'])
        registerStudent.delay(name, usn, dept.id, div.id, imgs)
        return redirect('studentregister')
    return render(request, 'student/register.html', context)

@login_required(login_url="/student/loginrequired/")
def classromm_list(request):
    context = {}
    context['classrooms'] = Classrooms.objects.all()
    return render(request, 'student/classrooms.html', context)


@login_required(login_url="/student/loginrequired/")
def success_page(request):
    return render(request, 'student/success.html')

@login_required(login_url="/student/loginrequired/")
def register_class(request):
    context = {}
    context['classrooms'] = Classrooms.objects.all()
    context['depts'] = Department.objects.all()
    context['divs'] = Division.objects.all()
    if request.method == "POST":
        cno = request.POST.get("cno")
        dept = request.POST.get("dept")
        div = request.POST.get("div")
        cip = request.POST.get("cip")

        classObj = Classrooms.objects.create(
            cno=cno,
            dept = Department.objects.get(name=dept),
            div = Division.objects.get(div_name=div),
            cip = cip
        )
        classObj.save()
        return redirect('classrooms')
    return render(request, 'student/registerclass.html', context)

def login_mandatory(request):
    return render(request, 'student/logrequire.html')